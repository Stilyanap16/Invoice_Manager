from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import datetime
import xlrd
import re
from database.database import get_data, insert_data, get_column_names, make_record_readable, INVOICE_NUMBER_REGEX

def import_from_excel(file_path) -> str:
    try:
        wb = load_workbook(file_path, data_only=False)
        sheet = wb.active

        max_col = len(get_column_names()[1:])
        max_row = sheet.max_row
        for row in range(max_row, 0, -1):
            if sheet.cell(row=row, column=2).value is not None:
                max_row = row
                break

        count = 0
        if max_row >= 2:
            for row in sheet.iter_rows(min_row=2, max_col=max_col, max_row=max_row, values_only=False):
                data = validate_data(row)
                if data:
                    insert_data(data)
                    count += 1

        result = "Импортирането е завършено. "
        if count == 0:
            result += "Няма правилни записи във файла."
        elif count == 1:
            result += "Добави се 1 правилен запис."
        else:
            result += f"Добавиха се {count} правилни записа."
        return result

    except InvalidFileException:
        return "Файлът не е валиден Excel файл."
    except FileNotFoundError:
        return "Файлът не е намерен."
    except:
        return "Грешка."

def export_to_excel() -> str:
    wb = Workbook()
    sheet = wb.active
    sheet.append(get_column_names()[1:])
    data = get_data()

    if data:
        for row in data:
            row = make_record_readable(list(row))
            sheet.append(row[1:])
        for cell in sheet["D"][1:]:
            cell.number_format = "0.00"
        for cell in sheet["E"][1:]:
            cell.number_format = "0.00"
        output_file = "exported_data.xlsx"
        wb.save(output_file)
        return f"Excel файлът {output_file} е попълнен успешно."
    return "В базата данни няма записи все още."

def validate_data(row: tuple) -> list:
    data = []
    try:
        date = row[0].value
        if isinstance(date, (datetime.date, datetime.datetime)):
            date = date.date()
        else:
            date = xlrd.xldate_as_datetime(row[0].value, 0).date()
        data.append(date.strftime("%Y-%m-%d"))
    except TypeError:
        return []

    if (row[1].value is None) or (row[1].data_type != 's'):
        return []
    invoice_number = row[1].value
    if re.fullmatch(INVOICE_NUMBER_REGEX, invoice_number) is None:
        return []
    data.append(invoice_number)

    if (row[2].value is None) or (row[2].data_type != 's'):
        return []
    client = row[2].value
    data.append(client)

    try:
        figure = float(row[3].value)
        if figure <= 0:
            return []
        data.append(figure)
    except (ValueError, TypeError):
        return []

    try:
        tax = float(row[4].value)
        if tax <= 0:
            raise ValueError
        data.append(tax)
    except (ValueError, TypeError):
        data.append(0.2 * float(row[3].value))

    if (row[5].value is None) or (row[5].data_type != 's'):
        return []
    category = row[5].value
    data.append(category)

    if row[6].value not in ('Да', 'Не'):
        return []
    is_paid = False
    if row[6].value == 'Да':
        is_paid = True
    data.append(is_paid)

    return data
